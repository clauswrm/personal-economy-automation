import openpyxl as excel


def findOpenRows(sheet, category, col="B"):
	"""
	Finds the range of open rows to write new transactions of a given category to.

	:param sheet: The sheet to search in.
	:param category: A string specifying the transaction type you want to find space for.
	:param col: The column ot look for category names in.
	:return: A list of the first open row in the given category and the last.
	:raises ValueError: If the category isn't found in the datasheet.
	"""
	category_start_row, category_end_row = 0, 0
	for cell in sheet[col]:
		if cell.value == category:
			category_start_row = cell.row
		elif cell.value == "Totalt" and category_start_row != 0:
			category_end_row = cell.row - 1
			break
	else:
		raise ValueError("Category " + category + " not found.")

	category_start_coord = chr(ord(col) + 1) + str(category_start_row)
	category_end_coord = chr(ord(col) + 1) + str(category_end_row)
	for cell in sheet[category_start_coord:category_end_coord]:
		if cell[0].value is None:
			return [cell[0].row, category_end_row]


def extractTransactionData(sheet):
	"""
	Extracts the transaction data from the given datasheet.

	:param sheet: A datasheet in the standard "Nordea" format.
	:return: A list containing each transaction as a tuple with the format (date, text, amount).
	"""
	valueColumns = (2, 6, 8, 10)  # All columns we want to extract information from
	transactions = []
	for r in range(sheet.max_row - 3, 10, -1):
		transaction = []
		for col in valueColumns:
			cell = sheet.cell(row=r, column=col)
			if cell.value is not None:
				transaction.append(cell.value)
		transactions.append(transaction)
	return transactions


def insertTransactionData(sheet, transactions):
	"""
	Inserts the transaction data into the specified datasheet.

	:param sheet: A datasheet in the standard "ClausRegnskap" format.
	:param transactions: A list of all transactions to be inserted to the sheet.
	:return: If the insertion of all transactions was successful or not.
	"""
	categories = ("Mat & forbruk", "Fest, gøy & snop", "Bolig og interiør", "Skole & matriell", "Diverse")
	categoryRows = {c: findOpenRows(sheet, c) for c in categories}

	for transaction in transactions:
		category = getCategory(transaction)
		if category == "manual":
			print("T:", transaction[1])
			while category not in categories:
				try:
					category = categories[int(input("Please name the category (0 - 4): "))]
				except (ValueError, IndexError) as e:
					print("Try again")

		startIndex, endIndex = "C" + str(categoryRows[category][0]), "E" + str(categoryRows[category][0])
		cells = sheet[startIndex:endIndex][0]  # Sheet slicing returns a tuple of a single tuple, therfore the [0]
		writeTransaction(transaction, cells)

		categoryRows[category][0] += 1
		if categoryRows[category][0] == categoryRows[category][1]:
			# Expand category and recalculate positions?
			return False
	else:
		return True


def writeTransaction(transaction, cells):
	""" Writes the transaction (date, text, amount) to the cells """
	for field, cell in zip(transaction, cells):
		cell.value = field


def getCategory(transaction):
	categories = {"Mat & forbruk": ["rema", "bunnpris", "kiwi", "coop", "extra", "joker", "hangaren", "sit",
									"petters pizza", "resturant", "kiosk"],
				  "Fest, gøy & snop": ["serveringsgjeng", "samfundet", "vinmonopolet", "reddkjellerne"],
				  "Bolig og interiør": ["ikea", "tef/bbl", "fra bendik schrøder", "lån ", "kraft"],
				  "Skole & matriell": ["akademika"],
				  "Diverse": ["atb ", "agathon", "ruter", "shell", "xxl", "yx ", "flytoget", "norwegian"]}

	text = transaction[1].lower()
	for category in categories:
		for word in categories[category]:
			if word in text:
				return category
	return "manual"


def numberToLetter(num):
	""" Transforms a column index to the corresponding excel column letter """
	return chr(num + ord("A") - 1)


def main():
	extractWB = excel.load_workbook('test.xlsx')
	insertWB = excel.load_workbook('testLagre.xlsx')
	extractSheet = extractWB.get_active_sheet()
	insertSheet = insertWB.get_active_sheet()
	l = extractTransactionData(extractSheet)
	print(l)
	b = insertTransactionData(insertSheet, l)
	print(b)
	insertWB.save('testLagre.xlsx')


if __name__ == '__main__':
	main()
